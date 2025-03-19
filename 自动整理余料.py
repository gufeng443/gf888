import time
import os
import shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import random
from datetime import datetime
from selenium.webdriver.chrome.service import Service
import threading

# 全局变量
browser_windows = {}
selected_browser = None
driver = None
last_excel_file = None
progress_bar = None
is_fetching = False


def start_chrome_with_debugging():
    global driver
    chrome_options = Options()
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument("--user-data-dir=C:\\Temp\\ChromeProfile")

    try:
        driver_path = r"E:\GP\chromedriver-win64\chromedriver-win64\chromedriver.exe"
        service = Service(executable_path=driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.get("https://web.whatsapp.com/")
        status_label.config(text="Chrome 已启动并打开 WhatsApp Web！")
    except Exception as e:
        status_label.config(text=f"启动失败: {str(e)}\n请检查：\n1. Selenium版本是否过时\n2. 驱动路径是否正确")


def get_browser_windows():
    global browser_windows, driver
    if driver is None:
        status_label.config(text="请先启动 Chrome！")
        return

    try:
        handles = driver.window_handles
        browser_windows = {handle: driver.title for handle in handles}
        browser_dropdown['values'] = list(browser_windows.values())
        status_label.config(text="成功获取浏览器窗口！")
    except Exception as e:
        status_label.config(text=f"获取窗口失败: {str(e)}")


def select_browser_window(event):
    global selected_browser, driver
    selected_window_title = browser_dropdown.get()
    for handle, title in browser_windows.items():
        if title == selected_window_title:
            selected_browser = handle
            driver.switch_to.window(selected_browser)
            status_label.config(text=f"已选择窗口: {title}")
            break


def slow_scroll(driver):
    scroll_pause_time = random.uniform(0.5, 1.5)
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(scroll_pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


def get_failed_group_names():
    global is_fetching
    if selected_browser is None:
        status_label.config(text="请先选择一个浏览器窗口！")
        return

    if is_fetching:
        return

    is_fetching = True
    threading.Thread(target=fetch_groups_thread).start()


def fetch_groups_thread():
    global is_fetching
    try:
        # 在主线程更新UI
        root.after(0, lambda: [
            status_label.config(text="正在获取群组数据..."),
            progress_bar.start()
        ])

        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                 "div.x1c4vz4f.xs83m0k.xdl72j9.x1g77sc7.x78zum5.xozqiw3.x1oa3qoh.x12fk4p8.xeuugli.x2lwn1j.x1nhvcw1.xdt5ytf.x1cy8zhl"))
        )
        slow_scroll(driver)
        elements = driver.find_elements(By.CSS_SELECTOR,
                                        "div.x1c4vz4f.xs83m0k.xdl72j9.x1g77sc7.x78zum5.xozqiw3.x1oa3qoh.x12fk4p8.xeuugli.x2lwn1j.x1nhvcw1.xdt5ytf.x1cy8zhl span.x193iq5w.xuxw1ft.xlyipyv.x6ikm8r.x10wlt62._ao3e")

        data = []
        total = len(elements)
        for i, element in enumerate(elements):
            data.append(element.text + "\n")
            time.sleep(random.uniform(1, 3))

            # 更新进度
            root.after(0, lambda i=i: (
                progress_bar.config(value=(i + 1) / total * 100),
                status_label.config(text=f"已获取 {i + 1}/{total} 个群组")
            ))

        root.after(0, lambda: (
            data_textbox.delete(1.0, tk.END),
            data_textbox.insert(tk.END, "".join(data)),
            update_number_box(),
            progress_bar.stop(),
            progress_bar.config(value=0)
        ))

    except Exception as e:
        root.after(0, lambda: [
            status_label.config(text=f"数据获取失败: {str(e)}"),
            progress_bar.stop()
        ])
    finally:
        is_fetching = False
        root.after(0, lambda: progress_bar.grid_remove())


def update_number_box():
    data = data_textbox.get(1.0, tk.END).strip().split("\n")
    numbers = []
    for line in data:
        if "-" in line:
            number = line.rsplit("-", 1)[-1].strip()
            if number.isdigit():
                numbers.append(number)

    delete_rows_entry.delete(0, tk.END)
    delete_rows_entry.insert(0, ", ".join(numbers))


def find_file(base_dir, file_name):
    for root, dirs, files in os.walk(base_dir):
        if file_name in files:
            return os.path.join(root, file_name)
    return None


def export_data():
    source_dir = r"C:\Users\Administrator\Desktop\料子1"
    target_base_dir = r"C:\Users\Administrator\Desktop\余料"

    data = data_textbox.get(1.0, tk.END).strip().split("\n")
    if not data:
        messagebox.showwarning("警告", "文本框中没有数据！")
        return

    if not os.path.exists(target_base_dir):
        os.makedirs(target_base_dir)

    for line in data:
        if "-" in line:
            folder_name = line.split("-", 1)[0].strip()
            target_dir = os.path.join(target_base_dir, folder_name)

            if not os.path.exists(target_dir):
                os.makedirs(target_dir)

            file_name = f"{line.strip()}.txt"
            file_path = find_file(source_dir, file_name)
            if not file_path:
                status_label.config(text=f"文件不存在: {file_name}")
                continue

            try:
                shutil.move(file_path, target_dir)
                status_label.config(text=f"文件已移动: {file_name} -> {target_dir}")
            except Exception as e:
                status_label.config(text=f"移动文件失败: {str(e)}")

    messagebox.showinfo("完成", "数据导出完成！")


def modify_excel(file_path, actual_pull_count, rows_to_delete):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        if rows_to_delete.strip():
            rows_to_remove = []
            rows_to_delete = set(map(int, rows_to_delete.split(',')))
            for row in range(2, sheet.max_row + 1):
                group_name = sheet.cell(row=row, column=5).value
                if group_name:
                    group_number = int(group_name.split('-')[-1])
                    if group_number in rows_to_delete:
                        rows_to_remove.append(row)
            for row in reversed(rows_to_remove):
                sheet.delete_rows(row)

        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        task_count = sheet.max_row - 3
        correct_task_count = task_count - 1
        sheet.cell(row=sheet.max_row, column=1).value = f"数据导出时间:{current_time}\n任务数:{correct_task_count}"

        rows = list(range(2, sheet.max_row - 2))
        group_count = len(rows)
        max_pullable = group_count * 4

        user_pull = int(actual_pull_count)
        if user_pull > max_pullable:
            messagebox.showerror("错误", f"输入人数超过最大可分配值（{group_count}个群组 × 4 = {max_pullable}）")
            return
        if user_pull < 0:
            messagebox.showerror("错误", "实际拉人数不能为负数")
            return

        per_group = user_pull // group_count
        remainder = user_pull % group_count
        allocations = [per_group] * group_count

        indexes = list(range(group_count))
        random.shuffle(indexes)
        for i in indexes[:remainder]:
            allocations[i] += 1

        total_member_count = 0
        for row, allocation in zip(rows, allocations):
            sheet.cell(row=row, column=2).value = allocation
            sheet.cell(row=row, column=3).value = 4 - allocation
            member_count = 2 + allocation
            sheet.cell(row=row, column=1).value = member_count
            total_member_count += member_count

        sheet.cell(row=sheet.max_row - 2, column=1).value = f"总人数: {total_member_count}"
        sheet.cell(row=sheet.max_row - 2, column=2).value = f"总拉人数: {user_pull}"

        workbook.save(file_path)
        messagebox.showinfo("成功", "文件修改成功！")
    except Exception as e:
        messagebox.showerror("错误", f"发生错误: {e}")


def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)


def start_execution():
    file_path = file_entry.get()
    actual_pull_count = actual_pull_entry.get()
    rows_to_delete = delete_rows_entry.get()

    if not file_path:
        messagebox.showerror("错误", "请选择Excel文件")
        return
    if not actual_pull_count.isdigit():
        messagebox.showerror("错误", "实际人数必须为数字")
        return

    modify_excel(file_path, actual_pull_count, rows_to_delete)


def update_calculator(*args):
    try:
        input_text = delete_rows_entry.get()
        count = len(input_text.split(',')) if input_text.strip() else 0
        result = (100 - count) * 3.7
        calculator_label.config(text=f"计算结果: (100 - {count}) * 3.7 = {result:.2f}")
    except Exception as e:
        calculator_label.config(text=f"计算错误: {str(e)}")


# 创建GUI界面
root = tk.Tk()
root.title("浏览器自动化与 Excel 修改工具")

main_frame = ttk.Frame(root, padding="10")
main_frame.pack(fill=tk.BOTH, expand=True)

# 浏览器控制部分
browser_frame = ttk.LabelFrame(main_frame, text="浏览器自动化", padding="10")
browser_frame.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)

ttk.Button(browser_frame, text="启动 Chrome", command=start_chrome_with_debugging).grid(row=0, column=0, padx=5, pady=5)
ttk.Button(browser_frame, text="获取窗口", command=get_browser_windows).grid(row=1, column=0, padx=5, pady=5)
browser_dropdown = ttk.Combobox(browser_frame, state="readonly", width=20)
browser_dropdown.grid(row=2, column=0, padx=5, pady=5)
browser_dropdown.bind("<<ComboboxSelected>>", select_browser_window)
ttk.Button(browser_frame, text="获取群组", command=get_failed_group_names).grid(row=3, column=0, padx=5, pady=5)

# 在browser_frame中添加进度条
progress_bar = ttk.Progressbar(browser_frame, orient='horizontal', length=200, mode='determinate')
progress_bar.grid(row=4, column=0, padx=5, pady=5, sticky='ew')
progress_bar.grid_remove()

# 数据显示部分
data_frame = ttk.LabelFrame(main_frame, text="数据展示", padding="10")
data_frame.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)

data_textbox = tk.Text(data_frame, height=10, width=35)
data_textbox.grid(row=0, column=0, padx=5, pady=5)
ttk.Button(data_frame, text="导出数据", command=export_data).grid(row=1, column=0, padx=5, pady=5)

# 删除控制部分
delete_frame = ttk.LabelFrame(main_frame, text="删除管理", padding="10")
delete_frame.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)

ttk.Label(delete_frame, text="删除行号（逗号分隔）:").grid(row=0, column=0, padx=5, pady=5)
delete_rows_entry = ttk.Entry(delete_frame, width=20)
delete_rows_entry.grid(row=1, column=0, padx=5, pady=5)
delete_rows_entry.bind("<KeyRelease>", update_calculator)

calculator_label = ttk.Label(delete_frame, text="计算结果: (100 - 0) × 3.7 = 370.00")
calculator_label.grid(row=2, column=0, padx=5, pady=5)

# Excel控制部分
excel_frame = ttk.LabelFrame(main_frame, text="Excel操作", padding="10")
excel_frame.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)

ttk.Label(excel_frame, text="Excel文件:").grid(row=0, column=0, padx=5, pady=5)
file_entry = ttk.Entry(excel_frame, width=20)
file_entry.grid(row=1, column=0, padx=5, pady=5)
ttk.Button(excel_frame, text="浏览", command=select_file).grid(row=1, column=1, padx=5, pady=5)

ttk.Label(excel_frame, text="实际人数:").grid(row=2, column=0, padx=5, pady=5)
actual_pull_entry = ttk.Entry(excel_frame, width=20)
actual_pull_entry.grid(row=3, column=0, padx=5, pady=5)

ttk.Button(excel_frame, text="执行修改", command=start_execution).grid(row=4, column=0, padx=5, pady=5)

# 状态栏
status_label = ttk.Label(main_frame, text="就绪")
status_label.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

root.mainloop()

if driver:
    driver.quit()