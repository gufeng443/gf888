import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import random
from datetime import datetime


def modify_excel(file_path, actual_pull_count, rows_to_delete):
    try:
        # 加载Excel文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 如果用户输入了要删除的行，执行删除操作
        if rows_to_delete.strip():
            rows_to_delete = set(map(int, rows_to_delete.split(',')))
            rows_to_remove = []
            for row in range(2, sheet.max_row + 1):
                group_name = sheet.cell(row=row, column=5).value
                if group_name:
                    group_number = int(group_name.split('-')[-1])
                    if group_number in rows_to_delete:
                        rows_to_remove.append(row)
            for row in reversed(rows_to_remove):
                sheet.delete_rows(row)

        # 更新任务数

        # 获取当前时间并格式化为 "YYYYMMDDHHMMSS" 格式
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")

        # 更新任务数
        task_count = sheet.max_row - 3  # 减去标题行、总人数行和空白行
        correct_task_count = task_count - 1  # 任务数需要再减去 1

        # 将数据导出时间和任务数写入单元格
        sheet.cell(row=sheet.max_row, column=1).value = f"数据导出时间:{current_time}\n任务数:{correct_task_count}"

        # 计算剩余群组的总人数和最大可拉人数
        rows = list(range(2, sheet.max_row - 2))
        total_member_count = 0
        max_pullable = 0
        for row in rows:
            member_count = sheet.cell(row=row, column=1).value
            if member_count is not None:
                total_member_count += member_count
                max_pullable += min(member_count, 4)  # 每个群组最多拉4人

        # 校验用户输入的实际拉人数
        user_pull = int(actual_pull_count)
        if user_pull > max_pullable:
            messagebox.showerror("错误", f"实际拉人数超过最大可分配值（{max_pullable}）")
            return
        if user_pull < 0:
            messagebox.showerror("错误", "实际拉人数不能为负数")
            return

        remaining_pull = user_pull
        random.shuffle(rows)  # 随机打乱行顺序

        # 第一阶段：初始分配（随机分配2-4人）
        for row in rows:
            member_count = sheet.cell(row=row, column=1).value
            if member_count is not None:
                max_pull = min(member_count, 4)
                pull = random.randint(2, max_pull) if max_pull >= 2 else 0
                pull = min(pull, remaining_pull)  # 确保不超过剩余拉人数
                sheet.cell(row=row, column=2).value = pull
                sheet.cell(row=row, column=3).value = member_count - pull
                remaining_pull -= pull

        # 第二阶段：补充分配（从小群组开始增加）
        if remaining_pull > 0:
            # 按当前拉人数从小到大排序
            sorted_rows = sorted(
                [(row, sheet.cell(row=row, column=2).value) for row in rows],
                key=lambda x: x[1]
            )
            for row, current_pull in sorted_rows:
                member_count = sheet.cell(row=row, column=1).value
                if member_count is None:
                    continue
                max_possible = min(member_count, 4)
                available = max_possible - current_pull
                if available <= 0:
                    continue
                add = min(available, remaining_pull)
                sheet.cell(row=row, column=2).value += add
                sheet.cell(row=row, column=3).value -= add
                remaining_pull -= add
                if remaining_pull == 0:
                    break

        # 最终校验
        total_pull = sum(sheet.cell(row=row, column=2).value for row in rows)
        if total_pull != user_pull:
            messagebox.showerror("错误", f"分配失败，总和为{total_pull}，预期{user_pull}")
            return

        # 更新总人数和总拉人数
        sheet.cell(row=sheet.max_row - 2, column=1).value = f"总人数: {total_member_count}"
        sheet.cell(row=sheet.max_row - 2, column=2).value = f"总拉人数: {user_pull}"

        # 保存文件
        workbook.save(file_path)
        messagebox.showinfo("成功", "文件修改成功！")
    except Exception as e:
        messagebox.showerror("错误", f"发生错误: {e}")


# 其余界面代码保持不变...
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, file_path)


def start_execution():
    file_path = file_entry.get()
    actual_pull_count = actual_pull_entry.get()
    rows_to_delete = delete_rows_entry.get()

    if not file_path:
        messagebox.showerror("错误", "请选择Excel文件")
        return
    if not actual_pull_count.isdigit():
        messagebox.showerror("错误", "实际人数必须为数字")
        return

    modify_excel(file_path, actual_pull_count, rows_to_delete)


# 创建用户界面
root = tk.Tk()
root.title("Excel 修改工具")

tk.Label(root, text="选择Excel文件:").grid(row=0, column=0, padx=5, pady=5)
file_entry = tk.Entry(root, width=50)
file_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="浏览", command=select_file).grid(row=0, column=2, padx=5, pady=5)

tk.Label(root, text="实际人数:").grid(row=1, column=0, padx=5, pady=5)
actual_pull_entry = tk.Entry(root, width=50)
actual_pull_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="清除指定数据行(逗号分隔):").grid(row=2, column=0, padx=5, pady=5)
delete_rows_entry = tk.Entry(root, width=50)
delete_rows_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Button(root, text="开始执行", command=start_execution).grid(row=3, column=1, padx=5, pady=5)

root.mainloop()