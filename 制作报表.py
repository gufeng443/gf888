import time
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def modify_excel_data(excel_path, input_people_count, group_name, group_link):
    """
    修改 Excel 文件中的数据
    :param excel_path: Excel 文件路径
    :param input_people_count: 输入的人数
    :param group_name: 群名称
    :param group_link: 群链接
    """
    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_path)
        ws = wb.active

        # 修改第一行第二列的数据（拉人数）
        ws.cell(row=2, column=1).value = input_people_count

        # 修改第二行第二列的数据（拉人数）
        ws.cell(row=2, column=2).value = input_people_count

        # 修改第三行第一列的数据（总人数）
        ws.cell(row=3, column=1).value = f"总人数：{input_people_count}"

        # 修改第三行第二列的数据（拉群人数）
        ws.cell(row=3, column=2).value = f"拉群人数：{input_people_count}"

        # 修改第二行第五列的群名称信息
        ws.cell(row=2, column=5).value = group_name

        # 修改第二行第四列的群链接信息
        ws.cell(row=2, column=4).value = group_link

        # 修改第五行第一列的时间信息
        current_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
        ws.cell(row=5, column=1).value = f"数据导出时间：{current_time}"

        # 保存修改后的 Excel 文件
        wb.save(excel_path)
        messagebox.showinfo("成功", "Excel 文件修改成功！")
    except Exception as e:
        messagebox.showerror("错误", f"修改 Excel 文件时发生错误：{e}")

def save_as_new_file(excel_path, new_filename):
    """
    将文件另存为新的文件名
    :param excel_path: 原文件路径
    :param new_filename: 新文件名
    """
    try:
        # 获取原文件所在目录
        directory = os.path.dirname(excel_path)
        # 构建新文件路径
        new_file_path = os.path.join(directory, new_filename)

        # 确保新文件名以 .xlsx 结尾
        if not new_file_path.endswith(".xlsx"):
            new_file_path += ".xlsx"

        # 加载原文件并另存为新文件
        wb = load_workbook(excel_path)
        wb.save(new_file_path)
        messagebox.showinfo("成功", f"文件已成功另存为：{new_file_path}")
    except Exception as e:
        messagebox.showerror("错误", f"另存为文件时发生错误：{e}")

def select_file():
    """选择 Excel 文件"""
    file_path = filedialog.askopenfilename(
        title="选择 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def start_modify():
    """开始修改 Excel 文件"""
    excel_path = entry_file_path.get()
    input_people_count = entry_people_count.get()
    group_name = entry_group_name.get()
    group_link = entry_group_link.get()

    if not excel_path:
        messagebox.showwarning("警告", "请选择 Excel 文件！")
        return

    if not input_people_count.isdigit():
        messagebox.showwarning("警告", "人数必须为数字！")
        return

    # 调用函数修改 Excel 数据
    modify_excel_data(excel_path, int(input_people_count), group_name, group_link)

def save_file():
    """保存为新文件"""
    excel_path = entry_file_path.get()
    new_filename = entry_save_as.get()

    if not excel_path:
        messagebox.showwarning("警告", "请先选择 Excel 文件！")
        return

    if not new_filename:
        messagebox.showwarning("警告", "请输入新文件名！")
        return

    # 调用函数另存为新文件
    save_as_new_file(excel_path, new_filename)

# 创建主窗口
root = tk.Tk()
root.title("Excel 数据修改工具")

# 设置窗口大小
root.geometry("450x350")

# 文件选择部分
label_file_path = tk.Label(root, text="Excel 文件路径：")
label_file_path.grid(row=0, column=0, padx=10, pady=10, sticky="w")

entry_file_path = tk.Entry(root, width=30)
entry_file_path.grid(row=0, column=1, padx=10, pady=10)

button_select_file = tk.Button(root, text="选择文件", command=select_file)
button_select_file.grid(row=0, column=2, padx=10, pady=10)

# 人数输入部分
label_people_count = tk.Label(root, text="人数：")
label_people_count.grid(row=1, column=0, padx=10, pady=10, sticky="w")

entry_people_count = tk.Entry(root, width=30)
entry_people_count.grid(row=1, column=1, padx=10, pady=10)

# 群名称输入部分
label_group_name = tk.Label(root, text="群名称：")
label_group_name.grid(row=2, column=0, padx=10, pady=10, sticky="w")

entry_group_name = tk.Entry(root, width=30)
entry_group_name.grid(row=2, column=1, padx=10, pady=10)

# 群链接输入部分
label_group_link = tk.Label(root, text="群链接：")
label_group_link.grid(row=3, column=0, padx=10, pady=10, sticky="w")

entry_group_link = tk.Entry(root, width=30)
entry_group_link.grid(row=3, column=1, padx=10, pady=10)

# 文件名输入部分
label_save_as = tk.Label(root, text="新文件名：")
label_save_as.grid(row=4, column=0, padx=10, pady=10, sticky="w")

entry_save_as = tk.Entry(root, width=30)
entry_save_as.grid(row=4, column=1, padx=10, pady=10)

# 开始修改按钮
button_start = tk.Button(root, text="开始修改", command=start_modify)
button_start.grid(row=5, column=1, padx=10, pady=10)

# 保存按钮
button_save = tk.Button(root, text="保存为新文件", command=save_file)
button_save.grid(row=6, column=1, padx=10, pady=10)

# 运行主循环
root.mainloop()