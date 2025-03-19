import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 定义输入文件路径
input_file_path = r'C:\Users\Administrator\Desktop\报表统计\交单.txt'

# 获取当前日期并格式化为 MM-DD
current_date_str = datetime.now().strftime("%m-%d")
output_file_name = f"统计{current_date_str}.xlsx"
output_file_path = os.path.join(r'C:\Users\Administrator\Desktop\报表统计', output_file_name)


# 尝试使用不同的编码读取文件
def read_file(file_path):
    encodings = ['gbk', 'utf-8']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                return file.readlines()
        except Exception as e:
            print(f"使用编码 {encoding} 读取文件失败: {e}")
    raise Exception("所有读取文件的尝试均失败。")


# 从输入文件读取行
lines = read_file(input_file_path)

# 提取数据
data = []
current_date = None

for line in lines:
    line = line.strip()
    if line:
        # 检查日期分割线
        date_match = re.match(r'(\d{1,2}\.\d{1,2})-+', line)
        if date_match:
            current_date = date_match.group(1).replace('.', '月') + '日'
            continue  # 跳过这一行，继续处理下一个

        # 提取人数
        match = re.search(r'-(\d+)人$', line)
        if match and current_date:
            person_count = match.group(1)
            data.append((current_date, line, person_count))

# 创建 DataFrame
df = pd.DataFrame(data, columns=['日期', '数据', '人数'])

# 将人数列转换为整数类型
df['人数'] = df['人数'].astype(int)

# 生成颜色列表，确保每个日期都有不同的颜色
color_list = [
    'FFDDFF',  # 浅紫色
    'FFDDDD',  # 浅红色
    'DDFFDD',  # 浅绿色
    'DDDDFF',  # 浅蓝色
    'FFFFDD',  # 浅黄色
    'DDFFFF',  # 浅青色
    'FFDDAA',  # 浅橙色
    'DDAAFF',  # 浅玫瑰色
    'FFAAAA',  # 浅桃色
    'AAFFAA',  # 浅草绿色
    'AAAAFF',  # 浅天蓝色
    'FFEEAA',  # 浅杏色
    'AAFFEE',  # 浅蓝绿色
    'FFCCAA',  # 浅肉色
    'CCFFCC',  # 浅草灰色
    'FFCCDD'  # 浅紫红色
]

# 创建颜色映射
date_colors = {}
unique_dates = df['日期'].unique()
for index, date in enumerate(unique_dates):
    if index < len(color_list):
        date_colors[date] = color_list[index]

# 尝试保存结果
try:
    with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='详细数据')

        # 获取工作表对象
        workbook = writer.book
        worksheet = writer.sheets['详细数据']

        # 添加颜色到每个日期行
        current_row = 2  # 数据从第二行开始
        for index, row in df.iterrows():
            if row['日期'] in date_colors:
                fill = PatternFill(start_color=date_colors[row['日期']], end_color=date_colors[row['日期']],
                                   fill_type='solid')
                worksheet.cell(row=current_row, column=1).fill = fill  # 日期列
                worksheet.cell(row=current_row, column=2).fill = fill  # 数据列
                worksheet.cell(row=current_row, column=3).fill = fill  # 人数列
            current_row += 1

except PermissionError:
    print("写入文件时出错: 请确保文件没有在其他程序中打开。")
except Exception as e:
    print(f"写入文件时出错: {e}")

print("数据统计完成！")
